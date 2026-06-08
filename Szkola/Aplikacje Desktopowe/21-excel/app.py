import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

#consts

EXCEL_FILE = "data.xlsx"
INCOME_CATEGORIES = ["Salary", "Freelance", "Interest", "Other"]
OUTCOME_CATEGORIES = ["Rent", "Groceries", "Utilities", "Entertainment", "Other"]

INCOME_COLOR = "99FF99"
OUTCOME_COLOR = "FF9999"
HEADER_COLOR = "CCCCCC"
ROW_COLOR = "ECF0F1"
ROW_PAIR_COLOR = 'ECF0F1'

def add_excel_transaction(path: str, record: dict):
    excel_init(path)
    wb = load_workbook(path)
    ws = wb.active
    new_row = ws.max_row + 1

    if new_row % 2:
        bg = PatternFill(fgColor=ROW_PAIR_COLOR)
    else:
        bg = PatternFill(fgColor=ROW_COLOR)

    sum_color = INCOME_COLOR if record["Type"] == "Income" else OUTCOME_COLOR
    value = [record['Date'], record['Type'], record['Category'], record['Amount'], record['Description']]

    for col, val in enumerate(value, start=1):
        cell = ws.cell(row=new_row, column=col, value=val)
        cell.fill = bg
        if col == 4:  # Amount column
            cell.font = Font(bold=True, color=sum_color)

    wb.save(path)

def generate_report(path: str, report_path: str):
    pass

def excel_init(path: str):
    if os.path.exists(path):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Type", "Category", "Amount", "Description"]
    ws.append(headers)

    fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=12)
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    widths = [14, 12, 20, 30, 16]

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    wb.save(path)

class BudgetManager():
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Budget Manager")
        self.root.geometry("800x600")
        self.root.resizable(False, False)

        excel_init(EXCEL_FILE)

        self.excel_path = EXCEL_FILE
        self.current_type = tk.StringVar(value="Income")
        self.selected_category = tk.StringVar(value=INCOME_CATEGORIES[0])
        self.entries = {}
        # build UI immediately
        self.build_gui()
        self.build_form()
        self.build_saldo()
        self.build_list()
        self.build_footer()
        self.refresh_list()

    def build_form(self):
        frame = tk.LabelFrame(self.left, text="Add Transaction", padx=8, pady=8, bg="#ECF0F1")
        frame.pack(fill=tk.X, padx=10, pady=10)

        fields = [
            ("Date (YYYY-MM-DD):", "date"),
            ("Type:", "type"),
            ("Category:", "category_menu"),
            ("Amount:", "amount"),
            ("Description:", "desc")
        ]

        self.variables: dict = {}

        for label, key in fields:
            tk.Label(frame, text=label, bg="#ECF0F1").grid(row=fields.index((label, key)), column=0, sticky=tk.W)

            if key == "type":
                self.variables[key] = tk.StringVar(value="Income")
                widget = ttk.Combobox(frame, values=["Income", "Outcome"], textvariable=self.variables[key], state="readonly")
            elif key == "category_menu":
                self.variables[key] = tk.StringVar(value=INCOME_CATEGORIES[0])
                widget = ttk.Combobox(frame, values=INCOME_CATEGORIES, textvariable=self.variables[key], state="readonly")
            else:
                widget = tk.Entry(frame)
                self.entries[key] = widget

            widget.pack(pady=4)
            tk.Button(frame, text="Add and save to excel", command=self.add_transaction, bg="#27AE60", fg="white").grid(row=5, column=0, columnspan=2, pady=8)
            

    
    def build_gui(self):
        header = tk.Frame(self.root, bg="#2C3E50", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="Budget Manager", bg="#2C3E50", fg="white", font=("Arial", 18)).pack(side=tk.LEFT, padx=20)

        file_frame = tk.Frame(self.root, bg="#ECF0F1", height=40)
        file_frame.pack(fill=tk.X)
        tk.Label(file_frame, text="Excel File:", bg="#ECF0F1", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)
        self.file_label = tk.Label(file_frame, text=self.excel_path, bg="#ECF0F1", font=("Arial", 12), fg="blue")
        self.file_label.pack(side=tk.LEFT, padx=10)

        main = tk.Frame(self.root, bg="#ECF0F1")
        main.pack(fill=tk.BOTH, expand=True)

        left = tk.Frame(main, bg="#ECF0F1", width=300)
        left.pack(side=tk.LEFT, fill=tk.Y)
        self.left = left

        right = tk.Frame(main, bg="#ECF0F1")
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        self.right = right

    def build_form(self):
        frame = tk.LabelFrame(self.left, text="Add Transaction", padx=8, pady=8, bg="#ECF0F1")
        frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(frame, text="Date (YYYY-MM-DD):", bg="#ECF0F1").grid(row=0, column=0, sticky=tk.W)
        date_entry = tk.Entry(frame)
        date_entry.grid(row=0, column=1, pady=4)

        tk.Label(frame, text="Type:", bg="#ECF0F1").grid(row=1, column=0, sticky=tk.W)
        rb_income = tk.Radiobutton(frame, text="Income", variable=self.current_type, value="Income", bg="#ECF0F1", command=self.change_type)
        rb_outcome = tk.Radiobutton(frame, text="Outcome", variable=self.current_type, value="Outcome", bg="#ECF0F1", command=self.change_type)
        rb_income.grid(row=1, column=1, sticky=tk.W)
        rb_outcome.grid(row=1, column=1, sticky=tk.E)

        tk.Label(frame, text="Category:", bg="#ECF0F1").grid(row=2, column=0, sticky=tk.W)
        category_menu = ttk.Combobox(frame, values=INCOME_CATEGORIES, textvariable=self.selected_category, state="readonly")
        category_menu.grid(row=2, column=1, pady=4)

        tk.Label(frame, text="Amount:", bg="#ECF0F1").grid(row=3, column=0, sticky=tk.W)
        amount_entry = tk.Entry(frame)
        amount_entry.grid(row=3, column=1, pady=4)

        tk.Label(frame, text="Description:", bg="#ECF0F1").grid(row=4, column=0, sticky=tk.W)
        desc_entry = tk.Entry(frame)
        desc_entry.grid(row=4, column=1, pady=4)

        add_btn = tk.Button(frame, text="Add", command=self.add_transaction, bg="#27AE60", fg="white")
        add_btn.grid(row=5, column=0, columnspan=2, pady=8)

        # store refs
        self.entries["date"] = date_entry
        self.entries["amount"] = amount_entry
        self.entries["desc"] = desc_entry
        self.entries["category_menu"] = category_menu

    def build_saldo(self):
        saldo_frame = tk.LabelFrame(self.right, text="Saldo", bg="#ECF0F1")
        saldo_frame.pack(fill=tk.X, padx=10, pady=10)
        self.saldo_label = tk.Label(saldo_frame, text="Balance: 0.00", bg="#ECF0F1", font=("Arial", 12))
        self.saldo_label.pack(anchor=tk.W, padx=8, pady=6)

    def build_list(self):
        list_frame = tk.Frame(self.right, bg="#ECF0F1")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        cols = ("Date", "Type", "Category", "Amount", "Description")
        tree = ttk.Treeview(list_frame, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, anchor=tk.W)

        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree = tree

    def build_footer(self):
        footer = tk.Frame(self.root, bg="#2C3E50", height=50)
        footer.pack(fill=tk.X, pady=(10, 0))
        buttons = [
            ("Refresh List", self.refresh_list, "#3498DB"),
            ("Change Type", self.change_type, "#3498DB"),
            ("Add Transaction", self.add_transaction, "#3498DB"),
            ("Export", self.export, "#3498DB"),
            ("Delete Selected", self.delete_selected, "#3498DB"),
            ("Choose File", self.choose_file, "#3498DB")
        ]
        for text, command, color in buttons:
            btn = tk.Button(footer, text=text, command=command, bg="#3498DB", fg="white", font=("Arial", 10), width=15)
            btn.pack(side=tk.LEFT, padx=10)

    def change_type(self):
        if self.current_type.get() == "Income":
            self.entries["category_menu"].config(values=INCOME_CATEGORIES)
            self.selected_category.set(INCOME_CATEGORIES[0])
        else:
            self.entries["category_menu"].config(values=OUTCOME_CATEGORIES)
            self.selected_category.set(OUTCOME_CATEGORIES[0])

    def add_transaction(self):
        date_text = self.entries["date"].get().strip() or datetime.now().strftime("%Y-%m-%d")
        try:
            # basic validation
            datetime.strptime(date_text, "%Y-%m-%d")
        except Exception:
            messagebox.showerror("Invalid date", "Enter date as YYYY-MM-DD")
            return

        ttype = self.current_type.get()
        category = self.selected_category.get()
        try:
            amount = float(self.entries["amount"].get())
        except Exception:
            messagebox.showerror("Invalid amount", "Enter a numeric amount")
            return
        desc = self.entries["desc"].get().strip()

        wb = load_workbook(self.excel_path)
        ws = wb.active
        ws.append([date_text, ttype, category, amount, desc])
        wb.save(self.excel_path)
        self.refresh_list()
        for k in ("date", "amount", "desc"):
            self.entries[k].delete(0, tk.END)

    def refresh_list(self):
        # clear tree
        for i in self.tree.get_children():
            self.tree.delete(i)

        wb = load_workbook(self.excel_path)
        ws = wb.active
        saldo = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            date, ttype, category, amount, desc = row
            amount = float(amount or 0)
            if ttype == "Income":
                saldo += amount
            else:
                saldo -= amount
            self.tree.insert("", tk.END, values=(date, ttype, category, amount, desc))

        self.saldo_label.config(text=f"Balance: {saldo:.2f}")

    def export(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if not path:
            return
        try:
            wb = load_workbook(self.excel_path)
            wb.save(path)
            messagebox.showinfo("Exported", f"Exported to {path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a row to delete")
            return
        vals = self.tree.item(sel[0], "values")
        # find and remove first matching row in excel
        wb = load_workbook(self.excel_path)
        ws = wb.active
        found = False
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            values = [cell.value for cell in row]
            # compare stringified values
            if tuple(str(v) for v in values[:5]) == tuple(str(v) for v in vals):
                ws.delete_rows(idx)
                found = True
                break
        if found:
            wb.save(self.excel_path)
            self.refresh_list()
        else:
            messagebox.showwarning("Not found", "Selected row not found in file")

    def choose_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")])
        if not path:
            return
        self.excel_path = path
        self.file_label.config(text=self.excel_path)
        excel_init(self.excel_path)
        self.refresh_list()




if __name__ == "__main__":
    root = tk.Tk()
    app = BudgetManager(root)
    root.mainloop()


